"""Справочник Excel, поиск строки, запись строки по шаблону."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string

from config import AppConfig, ChecklistField, PhotoSlot


def cell_to_str(v: Any) -> str:
    """Приведение значения ячейки к строке без артефактов типа '28.0'."""
    if pd.isna(v):
        return ""
    # numpy scalar -> python scalar
    if hasattr(v, "item"):
        try:
            v = v.item()
        except Exception:
            pass
    # 28.0 -> "28"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_registry(
    path: str | Path,
    sheet_name: int | str = 0,
    header_row: int = 1,
    start_row: int | None = None,
) -> pd.DataFrame:
    """Таблица реестра. header_row/start_row — 1-based строки Excel."""
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(str(p))
    # Явно читаем как текст, чтобы не появлялись значения вида "28.0"
    df = pd.read_excel(p, sheet_name=sheet_name, header=header_row - 1, dtype=str, keep_default_na=False)
    if start_row is None:
        return df
    # df начинается с (header_row+1)-й строки Excel как row 0
    offset = int(start_row) - (int(header_row) + 1)
    if offset > 0 and len(df) > 0:
        df = df.iloc[offset:].reset_index(drop=True)
    return df


def get_checklist_values_from_row(
    df: pd.DataFrame,
    row_idx: int,
    start_col_1based: int,
    n: int,
) -> list[str]:
    """Значения чек-листа из строки по позициям столбцов (1-based Excel)."""
    if row_idx < 0 or row_idx >= len(df) or n <= 0:
        return []
    c0 = start_col_1based - 1
    out: list[str] = []
    for j in range(n):
        ci = c0 + j
        if ci >= len(df.columns):
            out.append("")
            continue
        out.append(cell_to_str(df.iloc[row_idx, ci]))
    return out


def load_reference(path: str | Path, sheet_name: int | str | None = 0) -> pd.DataFrame:
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(str(p))
    # Явно читаем как текст, чтобы не появлялись значения вида "28.0"
    return pd.read_excel(p, sheet_name=sheet_name, dtype=str, keep_default_na=False)


def find_row_by_serial(
    df: pd.DataFrame,
    serial: str,
    serial_col: str,
) -> pd.Series | None:
    if serial_col not in df.columns:
        return None
    want = str(serial).strip()
    if not want:
        return None
    col = df[serial_col].map(cell_to_str)
    sub = df[col == want]
    if sub.empty:
        sub = df[col.str.lower() == want.lower()]
    if sub.empty:
        return None
    return sub.iloc[0]


def find_row_by_column(
    df: pd.DataFrame,
    value: str,
    column: str,
) -> pd.Series | None:
    """Поиск первой строки по точному совпадению в указанном столбце справочника."""
    if column not in df.columns:
        return None
    want = str(value).strip()
    if not want:
        return None
    col = df[column].map(cell_to_str)
    sub = df[col == want]
    if sub.empty:
        sub = df[col.str.lower() == want.lower()]
    if sub.empty:
        return None
    return sub.iloc[0]


def _normalize_cell(v: Any) -> Any:
    if pd.isna(v):
        return ""
    if hasattr(v, "item"):
        try:
            return v.item()
        except Exception:
            return v
    return v


def parse_checklist_start_column(spec: int | str) -> int:
    if isinstance(spec, int):
        return spec
    s = str(spec).strip().upper()
    if s.isdigit():
        return int(s)
    return column_index_from_string(s)


def _resolve_sheet(wb: openpyxl.Workbook, sheet: int | str) -> Any:
    if isinstance(sheet, int):
        return wb[wb.sheetnames[sheet]]
    return wb[str(sheet)]


def append_template_row(
    template_path: str | Path,
    cfg: AppConfig,
    row_values: dict[str, Any],
    checklist_values: list[Any],
    checklist_start: int | str,
    direct_columns: dict[int, Any] | None = None,
) -> None:
    """
    Дописывает строку в шаблон: значения по именам колонок из row_values,
    чек-лист подряд с колонки checklist_start (1-based).
    """
    p = Path(template_path)
    if not p.is_file():
        raise FileNotFoundError(str(p))

    wb = openpyxl.load_workbook(p)
    ws = _resolve_sheet(wb, cfg.template.sheet)
    hr = cfg.template.header_row
    start_idx = parse_checklist_start_column(checklist_start)

    if checklist_values:
        need_end = start_idx + len(checklist_values) - 1
    else:
        need_end = 0
    mc = ws.max_column or 1
    if need_end > 0:
        for c in range(mc + 1, need_end + 1):
            ws.cell(row=hr, column=c).value = f"Чеклист_{c - start_idx + 1}"

    headers: list[str | None] = []
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(row=hr, column=c).value
        headers.append(str(v).strip() if v is not None and str(v).strip() else None)

    start_row = int(getattr(cfg.template, "start_row", hr + 1))
    next_row = max(ws.max_row + 1, start_row)

    name_to_col: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h:
            name_to_col[h] = i + 1

    for key, val in row_values.items():
        col = name_to_col.get(str(key))
        if col is None:
            continue
        ws.cell(row=next_row, column=col).value = _normalize_cell(val)

    if direct_columns:
        for col, val in direct_columns.items():
            if not col or col < 1:
                continue
            ws.cell(row=next_row, column=int(col)).value = _normalize_cell(val)

    for i, val in enumerate(checklist_values):
        col = start_idx + i
        ws.cell(row=next_row, column=col).value = _normalize_cell(val)

    wb.save(p)


def delete_registry_data_row(path: str | Path, cfg: AppConfig, data_row_index: int) -> None:
    """Удаляет строку данных реестра: data_row_index — 0 для первой строки под заголовком."""
    if data_row_index < 0:
        raise ValueError("Некорректный индекс строки")
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(str(p))
    wb = openpyxl.load_workbook(p)
    ws = _resolve_sheet(wb, cfg.template.sheet)
    hr = cfg.template.header_row
    start_row = int(getattr(cfg.template, "start_row", hr + 1))
    excel_row = start_row + data_row_index
    if excel_row < start_row or excel_row > ws.max_row:
        raise ValueError("Строка вне диапазона листа")
    ws.delete_rows(excel_row, 1)
    wb.save(p)


def update_registry_data_row(
    path: str | Path,
    cfg: AppConfig,
    data_row_index: int,
    row_values: dict[str, Any],
    checklist_values: list[Any],
    checklist_start: int | str,
    direct_columns: dict[int, Any] | None = None,
) -> None:
    """Обновляет строку данных реестра: data_row_index — 0 для первой строки под заголовком."""
    if data_row_index < 0:
        raise ValueError("Некорректный индекс строки")
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(str(p))

    wb = openpyxl.load_workbook(p)
    ws = _resolve_sheet(wb, cfg.template.sheet)
    hr = cfg.template.header_row
    start_row = int(getattr(cfg.template, "start_row", hr + 1))
    start_idx = parse_checklist_start_column(checklist_start)
    excel_row = start_row + data_row_index
    if excel_row < start_row or excel_row > ws.max_row:
        raise ValueError("Строка вне диапазона листа")

    headers: list[str | None] = []
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(row=hr, column=c).value
        headers.append(str(v).strip() if v is not None and str(v).strip() else None)

    name_to_col: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h:
            name_to_col[h] = i + 1

    for key, val in row_values.items():
        col = name_to_col.get(str(key))
        if col is None:
            continue
        ws.cell(row=excel_row, column=col).value = _normalize_cell(val)

    if direct_columns:
        for col, val in direct_columns.items():
            if not col or col < 1:
                continue
            ws.cell(row=excel_row, column=int(col)).value = _normalize_cell(val)

    for i, val in enumerate(checklist_values):
        col = start_idx + i
        ws.cell(row=excel_row, column=col).value = _normalize_cell(val)

    wb.save(p)


def build_row_values_from_lookup(
    cfg: AppConfig,
    matched: pd.Series | None,
    serial: str,
    template_serial_column: str,
    field_entries: dict[str, str],
    semantic_values: dict[str, str],
    equipment_type: str,
    photo_filenames: dict[int, str],
    photo_slots: list[PhotoSlot],
) -> dict[str, Any]:
    """Собирает dict колонка шаблона -> значение."""
    out: dict[str, Any] = {}

    out[template_serial_column] = serial

    for template_col, ref_col in cfg.column_mapping.items():
        if ref_col not in (matched.index if matched is not None else []):
            continue
        if matched is None:
            continue
        val = matched.get(ref_col, "")
        out[template_col] = _normalize_cell(val)

    for k, v in field_entries.items():
        out[k] = v

    for sf in cfg.semantic:
        # semantic можно писать либо по имени колонки шаблона, либо по явной excel_column (см. direct_columns)
        if getattr(sf, "excel_column", None):
            continue
        out[sf.template_column] = semantic_values.get(sf.key, "")

    out[cfg.equipment_type_column] = equipment_type

    for ps in photo_slots:
        fn = photo_filenames.get(ps.slot, "")
        if cfg.photo_cell_mode == "fullpath" and fn:
            out[ps.template_column] = str(Path(fn).resolve())
        else:
            out[ps.template_column] = Path(fn).name if fn else ""

    return out


def collect_checklist_values(fields: list[ChecklistField], widget_getters: dict[str, Any]) -> list[Any]:
    vals = []
    for f in fields:
        g = widget_getters.get(f.id)
        if g is None:
            vals.append("")
            continue
        if callable(g):
            vals.append(g())
        else:
            vals.append(g)
    return vals
